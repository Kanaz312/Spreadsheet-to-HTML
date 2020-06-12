import openpyxl
from sys import argv
import winsound
import re


# The use of spaces before tags is intentional to have the final output be properly spaced out

# The CSS of of this site overwrites any inline image width settings, and doesn't allow any css/js scripts or the script/style tag/attribute
# To make all images look equal, crop them by hand to be square, then resize to 240px on Drupal

class Table():
    """Used to create the html code of a table with project entries for a given college"""
    def __init__(self, college):
        self.college = college
        self.empty = True
        # this is contains all of the lines to be added to the output file for a given table
        self.tags_list = [f"""<h1 class="college">{self.college}</h1>""","""<hr>""","""<table border="1" cellpadding="1" cellspacing="1" class="table_noStyle" id="table_styles">""", """    <thead>""", """        <tr>""", """            <th scope = "col">&nbsp;</th>""", """            <th scope = "col">&nbsp;</th>""", """            <th scope = "col">&nbsp;</th>""",
        """        </tr>""", """    </thead>""", """    <tbody>"""]
        # each inner list holds the info for the three projects in a given row
        self.entries= [[]]
        self.latest_entry = 0

    def __eq__(self, other):
        return self.college == other

    def add_title(self, title):
        """Adds an entry to a row of titles"""
        self.tags_list.append(f"""            <td width="270"><h2>{title}</h2>""")

    def add_rest(self, entry):
        """Add an entry to a row of info complete with embedded image, description of project, and pdf link"""
        title = entry[0]
        image = entry[1]
        descr = description_formattting(entry[2])
        pdf = entry[3]
        call_type = call_type_formatting(entry[4])
        department = entry[5]
        submitter = entry[6]
        prof = entry[7]
        if not image is None:
            self.tags_list.append(f"""            <td><img alt="{title}" src="this would be the link to the folder where this image is located/{image}"/><br />""")
            self.tags_list.append(f"""            <br /><b>Associated Professor</b>: {prof}<br />""")
        else:
            self.tags_list.append(f"""            <td><img alt="{title}" src="this would be the link to the template image"/><br />""")
            self.tags_list.append(f"""            <b>Associated Professor</b>: {prof}<br />""")
        self.tags_list.append(f"""            <b>Department</b>: {department}<br />""")
        self.tags_list.append(f"""            <b>CALL type</b>: {call_type}<br />""")
        self.tags_list.append(f"""            <b>Submitted by</b>: {submitter}<br />""")
        self.tags_list.append(f"""            <br />{descr}<br />""")
        if not pdf is None:
            self.tags_list.append(f"""            <a href="this would be the link to the folder where this pdf is located/{pdf}" target="_blank">More info</a><br /><br /><br />""")
        self.tags_list.append(f"""            </td>""")

    def add_title_row(self):
        """Starts the next row of project titles"""
        #use valign="bottom" to set the title alignment to bottom (doesn't look as good)
        self.tags_list.append("""        <tr valign="top">""")

    def end_row(self):
        """Ends the most recent row"""
        self.tags_list.append("""        </tr>""")

    def add_info_row(self):
        """Closes the current row and starts the next one"""
        self.tags_list.append("""        </tr>""")
        self.tags_list.append("""        <tr valign="top">""")
    
    def turn_to_html(self):
        """Takes the list of entries, converts it into an html table, turns the list of tags into a single string"""
        for row in self.entries:
            # create the row of titles
            self.add_title_row()
            for index in range(len(row)):
                self.add_title(row[index][0])
            # create the row of info for the corresponding titles
            self.add_info_row()
            for index in range(len(row)):
                self.add_rest(row[index])
            self.end_row()
        # close the table
        self.tags_list.append("""    </tbody>""")
        self.tags_list.append("""</table>\n""")
        # convert the list into a single string ready to write into the output file
        self.tags_list = "\n".join(self.tags_list)

    def add_entry(self, row):
        """Adds an entry of [project title, image, description, pdf, professor name]"""
        if len(self.entries[self.latest_entry]) == 3:
            self.entries.append([])
            self.latest_entry += 1
        for cell in [1, 3, 5, 6, 7, 8]:
            if row[cell].value is None:
                winsound.PlaySound("SystemHand", winsound.SND_ALIAS)
                raise ValueError(f"\nCELL EMPTY ERROR: {row[cell]}")
        self.entries[self.latest_entry].append([row[1].value, row[2].value, row[3].value, row[4].value, row[5].value, row[6].value, row[7].value, row[8].value])


def main(args):
    """Takes in the arguments from command line. Converts the given .xlsx sheet 
    into an html table written into the give .txt file"""
    args = try_to_open(args)
    wb = openpyxl.load_workbook(args[1])
    sheet = wb.active
    colleges = [Table("CAFES"), Table("CAED"), Table("CENG"), Table("CLA"), Table("COSAM"), Table("OCOB")]
    num_rows = sheet.max_row
    read_rows_to_table(sheet, num_rows, colleges)
    txt_name = get_txt_name(args)
    txt = open(txt_name, "w", encoding="utf-8")
    # write the header into the output txt file
    header = open("header.txt", "r")
    header_lines = header.readlines()
    for line in header_lines:
        txt.write(line)
    # write each college's table into the output txt file
    for college in colleges:
        if not college.empty:
            txt.write(college.tags_list)
    txt.close()
    print(f"\n{txt_name} was written")

def try_to_open(args):
    """Confirms that the .xlsx sheet is valid. Asks for an input .xlsx and output .txt if ones were not given."""
    try:
        wb = openpyxl.load_workbook(args[1])
        wb.close()
        return args
    except FileNotFoundError:
        winsound.PlaySound("SystemHand", winsound.SND_ALIAS)
        args[1] = input("\n\n[ERROR]: SPREADSHEET DOES NOT EXIST\nPlease re-enter the name of the spreadsheet [ex. Survey.xlsx]")
        return try_to_open(args)
    except openpyxl.utils.exceptions.InvalidFileException:
        winsound.PlaySound("SystemHand", winsound.SND_ALIAS)
        args[1] = input("\n\n[ERROR]: SPREADSHEET FILE IS NOT .XLSX\nPlease enter a new file (make sure to include .xlsx)[ex. Survey.xlsx]")
        return try_to_open(args)
    # ask for a sheet name and output file name if not given one
    except IndexError:
        args = ["Sheet_to_txt_list.py"]
        winsound.PlaySound("SystemHand", winsound.SND_ALIAS)
        args.append(input("Please type the name of the Excell Sheet to read [ex. Survey.xlsx]:\n"))
        args.append(input("Please type the name you want for the new file [ex. website.txt]: \n"))
        return try_to_open(args)

def read_rows_to_table(sheet, num_rows, colleges):
    """Reads the sheet from rows 2 to the end, adding each row to the entries of the corresponding college table.
    Then converts each college's tags into a single string html."""
    for row in sheet.iter_rows(min_row = 2, max_col=9, max_row= num_rows):
        # convert the image name to url format if the cell isn't empty
        if not row[2].value is None:
            row[2].value = convert_file_to_url(row[2].value)
        # convert the pdf name to url format if the cell isn't empty
        if not row[4].value is None:    
            row[4].value = convert_file_to_url(row[4].value)
        # add the entry to its corresponding college
        for college in colleges:
            if college == row[0].value:
                college.add_entry(row)
                college.empty = False
    # convert the entries to their html format
    for college in colleges:
        if not college.empty:
            college.turn_to_html()

def get_txt_name(args):
    """Confirms the name of the .txt file. If it already exists, ask for either a new name or permission to overwrite"""
    try:
        txt_name = args[2]
        if txt_name[len(txt_name) - 4::] != ".txt": txt_name += ".txt"
    except IndexError:
        winsound.PlaySound("SystemExlamation", winsound.SND_ALIAS)
        print("\nNext time, please put in an txt file name as the thrid argument.")
        txt_name = input("\nType the name of the txt file you wish to create [e.x. website.txt]:\n")
    overwrite = False
    same = True
    #continue to ask for a new file name if the file already exists or given permission to overwrite
    while same and not overwrite:
        if txt_name[len(txt_name) - 4::] != ".txt": txt_name += ".txt"
        try:
            test = open(txt_name, "r")
            test.close()
            winsound.PlaySound("SystemExlamation", winsound.SND_ALIAS)
            new_name = input(f"\n{txt_name} already exists, please type a different name for the new file. Or type \"overwrite\" to overwrite the existing file:\n")
            if new_name == "overwrite":
                overwrite = True
            else:
                txt_name = new_name
        except FileNotFoundError:
            same = False
    return txt_name

def convert_file_to_url(file_name):
    """Formats file names for url use"""
    file_name = file_name.replace(" ", "%20")
    file_name = file_name.replace("+", "%2B")
    return file_name

def call_type_formatting(call_type):
    """Turns the CALL type cell into a more readable format for the website
    (removes ending ";", cuts out the extra words part of other type)"""
    call_type = re.sub(r"&", r"and", call_type)
    call_type = re.sub(r";(?!.)", r"", call_type)
    call_type = re.sub(r" \(e\.g\., arts and culture or technology\)", r"", call_type)
    call_type = re.sub(r";", r"; ", call_type)
    return call_type

def description_formattting(descr):
    """Replaces some of the special characters from the descriptions
    and prevents issues with the "&" symbol and html"""
    descr = re.sub(r"&", r"and", descr)
    descr = re.sub(r"”|“", r"\"", descr)
    descr = re.sub(r"‛|’", r"\'", descr)
    descr = re.sub(r"‚", r",", descr)
    return descr





if __name__ == "__main__":
    main(argv)